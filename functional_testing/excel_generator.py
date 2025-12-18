"""
Excel Report Generator for Functional Testing
Generates Test_cases.xlsx and Bug_reports.xlsx
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generate Excel reports for test cases and bugs"""
    
    def __init__(self, output_dir="reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.not_run_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def _apply_header_style(self, ws, row=1, start_col=1, end_col=15):
        """Apply header styles"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
            
    def _apply_cell_style(self, ws, row, start_col=1, end_col=15):
        """Apply cell styles"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = self.border
            
    def _auto_adjust_column_width(self, ws, min_width=10, max_width=50):
        """Auto adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def generate_test_cases_excel(self, test_results, filename="Test_cases.xlsx"):
        """Generate Test_cases.xlsx"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Headers matching the provided template
        headers = [
            "Test Case ID",
            "Module",
            "Feature",
            "Testing Technique",
            "Test Case Description",
            "Precondition",
            "Test Steps",
            "Test Data",
            "Expected Result",
            "Actual Result",
            "Status",
            "Priority",
            "Execution Time",
            "Executed Date",
            "Executed By"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, 1, len(headers))
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Write test results
        for row_idx, result in enumerate(test_results, 2):
            ws.cell(row=row_idx, column=1, value=result.get('test_case_id', ''))
            ws.cell(row=row_idx, column=2, value=result.get('module', ''))
            ws.cell(row=row_idx, column=3, value=result.get('feature', ''))
            ws.cell(row=row_idx, column=4, value=result.get('technique', ''))
            ws.cell(row=row_idx, column=5, value=result.get('test_case', ''))
            ws.cell(row=row_idx, column=6, value=result.get('precondition', ''))
            ws.cell(row=row_idx, column=7, value=result.get('steps', ''))
            ws.cell(row=row_idx, column=8, value=result.get('test_data', ''))
            ws.cell(row=row_idx, column=9, value=result.get('expected_result', ''))
            ws.cell(row=row_idx, column=10, value=result.get('actual_result', ''))
            ws.cell(row=row_idx, column=11, value=result.get('status', ''))
            ws.cell(row=row_idx, column=12, value=result.get('priority', ''))
            ws.cell(row=row_idx, column=13, value=result.get('execution_time', ''))
            ws.cell(row=row_idx, column=14, value=result.get('executed_date', ''))
            ws.cell(row=row_idx, column=15, value=result.get('executed_by', ''))
            
            # Apply cell styles
            self._apply_cell_style(ws, row_idx, 1, len(headers))
            
            # Apply status color
            status_cell = ws.cell(row=row_idx, column=11)
            if result.get('status') == 'PASS':
                status_cell.fill = self.pass_fill
            elif result.get('status') == 'FAIL':
                status_cell.fill = self.fail_fill
            elif result.get('status') == 'NOT RUN':
                status_cell.fill = self.not_run_fill
                
        # Auto adjust columns
        self._auto_adjust_column_width(ws)
        
        # Add summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.cell(row=1, column=1, value="Test Execution Summary")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        total = len(test_results)
        passed = sum(1 for r in test_results if r.get('status') == 'PASS')
        failed = sum(1 for r in test_results if r.get('status') == 'FAIL')
        not_run = sum(1 for r in test_results if r.get('status') == 'NOT RUN')
        
        summary_data = [
            ("Total Test Cases", total),
            ("Passed", passed),
            ("Failed", failed),
            ("Not Run", not_run),
            ("Pass Rate", f"{(passed/total*100):.1f}%" if total > 0 else "0%"),
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
            
        # Technique breakdown
        ws_summary.cell(row=10, column=1, value="Breakdown by Technique")
        ws_summary.cell(row=10, column=1).font = Font(bold=True, size=12)
        
        techniques = {}
        for r in test_results:
            tech = r.get('technique', 'Unknown')
            if tech not in techniques:
                techniques[tech] = {'total': 0, 'pass': 0, 'fail': 0}
            techniques[tech]['total'] += 1
            if r.get('status') == 'PASS':
                techniques[tech]['pass'] += 1
            elif r.get('status') == 'FAIL':
                techniques[tech]['fail'] += 1
                
        ws_summary.cell(row=11, column=1, value="Technique")
        ws_summary.cell(row=11, column=2, value="Total")
        ws_summary.cell(row=11, column=3, value="Pass")
        ws_summary.cell(row=11, column=4, value="Fail")
        
        for row_idx, (tech, counts) in enumerate(techniques.items(), 12):
            ws_summary.cell(row=row_idx, column=1, value=tech)
            ws_summary.cell(row=row_idx, column=2, value=counts['total'])
            ws_summary.cell(row=row_idx, column=3, value=counts['pass'])
            ws_summary.cell(row=row_idx, column=4, value=counts['fail'])
            
        self._auto_adjust_column_width(ws_summary)
        
        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        print(f"Test cases exported to: {filepath}")
        return filepath
        
    def generate_bug_reports_excel(self, bugs, filename="Bug_reports.xlsx"):
        """Generate Bug_reports.xlsx"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bug Reports"
        
        # Headers matching the provided template
        headers = [
            "Bug ID",
            "Test Case ID",
            "Summary",
            "Description",
            "Severity",
            "Priority",
            "Module",
            "Feature",
            "Steps to Reproduce",
            "Expected Result",
            "Actual Result",
            "Status",
            "Assignee",
            "Found Date",
            "Fixed Date",
            "Verified By"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, 1, len(headers))
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Severity colors
        severity_fills = {
            'Critical': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'High': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'Medium': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'Low': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        }
        
        # Write bugs
        for row_idx, bug in enumerate(bugs, 2):
            bug_id = f"BUG_{row_idx-1:03d}"
            
            ws.cell(row=row_idx, column=1, value=bug_id)
            ws.cell(row=row_idx, column=2, value=bug.get('test_case_id', ''))
            ws.cell(row=row_idx, column=3, value=bug.get('summary', ''))
            ws.cell(row=row_idx, column=4, value=bug.get('description', ''))
            ws.cell(row=row_idx, column=5, value=bug.get('severity', 'Medium'))
            ws.cell(row=row_idx, column=6, value=bug.get('severity', 'Medium'))  # Priority = Severity
            ws.cell(row=row_idx, column=7, value=bug.get('module', ''))
            ws.cell(row=row_idx, column=8, value=bug.get('feature', ''))
            ws.cell(row=row_idx, column=9, value=bug.get('steps_to_reproduce', ''))
            ws.cell(row=row_idx, column=10, value=bug.get('expected', ''))
            ws.cell(row=row_idx, column=11, value=bug.get('actual', ''))
            ws.cell(row=row_idx, column=12, value="Open")
            ws.cell(row=row_idx, column=13, value="")  # Assignee
            ws.cell(row=row_idx, column=14, value=bug.get('found_date', ''))
            ws.cell(row=row_idx, column=15, value="")  # Fixed Date
            ws.cell(row=row_idx, column=16, value="")  # Verified By
            
            # Apply cell styles
            self._apply_cell_style(ws, row_idx, 1, len(headers))
            
            # Apply severity color
            severity = bug.get('severity', 'Medium')
            if severity in severity_fills:
                ws.cell(row=row_idx, column=5).fill = severity_fills[severity]
                
        # Add summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.cell(row=1, column=1, value="Bug Report Summary")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Count by severity
        severity_counts = {}
        for bug in bugs:
            sev = bug.get('severity', 'Medium')
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
            
        summary_data = [
            ("Total Bugs", len(bugs)),
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
            
        ws_summary.cell(row=6, column=1, value="Bugs by Severity")
        ws_summary.cell(row=6, column=1).font = Font(bold=True, size=12)
        
        for row_idx, (sev, count) in enumerate(severity_counts.items(), 7):
            ws_summary.cell(row=row_idx, column=1, value=sev)
            ws_summary.cell(row=row_idx, column=2, value=count)
            
        # Count by module
        module_counts = {}
        for bug in bugs:
            mod = bug.get('module', 'Unknown')
            module_counts[mod] = module_counts.get(mod, 0) + 1
            
        row_offset = 8 + len(severity_counts)
        ws_summary.cell(row=row_offset, column=1, value="Bugs by Module")
        ws_summary.cell(row=row_offset, column=1).font = Font(bold=True, size=12)
        
        for row_idx, (mod, count) in enumerate(module_counts.items(), row_offset + 1):
            ws_summary.cell(row=row_idx, column=1, value=mod)
            ws_summary.cell(row=row_idx, column=2, value=count)
            
        self._auto_adjust_column_width(ws)
        self._auto_adjust_column_width(ws_summary)
        
        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        print(f"Bug reports exported to: {filepath}")
        return filepath
        
    def generate_from_definition_only(self, test_cases, filename="Test_cases.xlsx"):
        """Generate Excel from test case definitions only (without execution)"""
        # Convert definition format to result format
        results = []
        for tc in test_cases:
            results.append({
                'test_case_id': tc['id'],
                'module': tc['module'],
                'feature': tc['feature'],
                'technique': tc['technique'],
                'test_case': tc['test_case'],
                'precondition': tc['precondition'],
                'steps': tc['steps'],
                'test_data': tc['test_data'],
                'expected_result': tc['expected_result'],
                'actual_result': '',
                'status': 'NOT RUN',
                'priority': tc.get('priority', 'Medium'),
                'execution_time': '',
                'executed_date': '',
                'executed_by': ''
            })
        return self.generate_test_cases_excel(results, filename)


if __name__ == "__main__":
    from test_cases_definition import ALL_TEST_CASES
    
    generator = ExcelReportGenerator(output_dir="reports")
    
    # Generate test cases without execution
    generator.generate_from_definition_only(ALL_TEST_CASES)
    
    # Generate empty bug report template
    generator.generate_bug_reports_excel([])
    
    print("\nExcel files generated successfully!")

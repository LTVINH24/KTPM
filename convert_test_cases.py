"""
Script chuyển đổi dữ liệu từ Test_cases.xlsx và Bug_reports.xlsx 
sang format Test Cases template-v1.1.xlsx và Bug_Report_Template.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from collections import OrderedDict


def convert_test_cases():
    """Chuyển đổi Test Cases"""
    # Đọc file nguồn (đường dẫn đúng)
    source_file = 'functional_testing/reports/Test_cases.xlsx'
    print(f'📖 Đang đọc file nguồn: {source_file}')
    wb_source = openpyxl.load_workbook(source_file)
    ws_all = wb_source['Test Cases']

    # Tạo workbook mới dựa trên template
    dest_template = 'docs/Test Cases template-v1.1.xlsx'
    print(f'📋 Đang đọc template: {dest_template}')
    wb_dest = openpyxl.load_workbook(dest_template)

    # --- CHUẨN BỊ DỮ LIỆU ---
    # Cột file nguồn: Test Case ID, Module, Feature, Testing Technique, Test Case Description, 
    # Precondition, Test Steps, Test Data, Expected Result, Actual Result, Status, Priority, 
    # Execution Time, Executed Date, Executed By
    test_cases = []
    for row in ws_all.iter_rows(min_row=2, values_only=True):
        if row[0]:
            test_cases.append({
                'tc_id': row[0],           # Test Case ID
                'module': row[1],          # Module
                'feature': row[2],         # Feature
                'technique': row[3],       # Testing Technique
                'test_name': row[4],       # Test Case Description
                'precondition': row[5],    # Precondition
                'test_steps': row[6],      # Test Steps
                'test_data': row[7],       # Test Data
                'expected': row[8],        # Expected Result
                'actual': row[9],          # Actual Result
                'status': row[10],         # Status
                'priority': row[11],       # Priority
                'execution_time': row[12], # Execution Time
                'executed_date': row[13],  # Executed Date
                'executed_by': row[14],    # Executed By
            })

    print(f'✅ Đọc được {len(test_cases)} test cases')

    # Nhóm theo Module và Feature
    features = OrderedDict()
    for tc in test_cases:
        module = tc['module']
        feature = tc['feature']
        key = f'{module} - {feature}'
        if key not in features:
            features[key] = []
        features[key].append(tc)

    print(f'📂 Nhóm thành {len(features)} features:')
    for f, tcs in features.items():
        print(f'   - {f}: {len(tcs)} test cases')

    # --- XỬ LÝ SHEET FEATURES ---
    ws_features = wb_dest['Features']
    # Xóa dữ liệu mẫu cũ (giữ header)
    ws_features.delete_rows(6, ws_features.max_row - 5)

    # Cập nhật thông tin project
    ws_features['A1'] = 'Project: OrangeHRM'
    ws_features['A2'] = 'Team: KTPM'

    # Thêm features
    feature_row = 6
    feature_id_map = {}
    for idx, (feature_name, tcs) in enumerate(features.items(), 1):
        feature_id = f'UC{idx:02d}'
        feature_id_map[feature_name] = feature_id
        
        ws_features[f'A{feature_row}'] = feature_id
        ws_features[f'B{feature_row}'] = feature_name
        ws_features[f'C{feature_row}'] = f'{len(tcs)} test cases - Technique: {tcs[0]["technique"]}'
        feature_row += 1

    print('✅ Đã tạo sheet Features')

    # --- XỬ LÝ SHEET TEST CASES ---
    ws_testcases = wb_dest['Test cases']

    # Styles
    feature_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Xóa dữ liệu mẫu cũ (giữ header row 1)
    ws_testcases.delete_rows(2, ws_testcases.max_row)

    # Thêm test cases theo format template
    current_row = 2
    for feature_name, tcs in features.items():
        feature_id = feature_id_map[feature_name]
        
        # Feature header row
        ws_testcases[f'A{current_row}'] = f'Feature {feature_id[-2:]}: {feature_name}'
        ws_testcases[f'A{current_row}'].font = Font(bold=True)
        for col in range(1, 11):
            cell = ws_testcases.cell(row=current_row, column=col)
            cell.fill = feature_fill
        current_row += 1
        
        # Test cases
        for tc_idx, tc in enumerate(tcs, 1):
            tc_id_short = f'TC{tc_idx:02d}'
            
            # Chuyển đổi status
            status_map = {'PASSED': 'Pass', 'FAILED': 'Fail', 'NOT RUN': 'N/A', 'Passed': 'Pass', 'Failed': 'Fail'}
            status = status_map.get(tc['status'], tc['status'])
            
            # Xử lý precondition + test_data + priority vào remark
            remark_parts = []
            if tc['precondition']:
                remark_parts.append(f"Precondition: {tc['precondition']}")
            if tc['test_data']:
                remark_parts.append(f"Test Data: {tc['test_data']}")
            if tc['priority']:
                remark_parts.append(f"Priority: {tc['priority']}")
            if tc['technique']:
                remark_parts.append(f"Technique: {tc['technique']}")
            remark = ' | '.join(remark_parts) if remark_parts else None
            
            # Ghi dữ liệu
            ws_testcases[f'A{current_row}'] = feature_id
            ws_testcases[f'B{current_row}'] = tc_id_short
            ws_testcases[f'C{current_row}'] = tc['test_name']
            ws_testcases[f'D{current_row}'] = tc['test_steps']
            ws_testcases[f'E{current_row}'] = tc['expected']
            ws_testcases[f'F{current_row}'] = tc['actual']
            ws_testcases[f'G{current_row}'] = status
            ws_testcases[f'H{current_row}'] = tc['executed_by']
            ws_testcases[f'I{current_row}'] = tc['executed_date']
            ws_testcases[f'J{current_row}'] = remark
            
            # Wrap text
            for col in range(1, 11):
                cell = ws_testcases.cell(row=current_row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
            
            current_row += 1

    # Điều chỉnh độ rộng cột
    ws_testcases.column_dimensions['A'].width = 12
    ws_testcases.column_dimensions['B'].width = 12
    ws_testcases.column_dimensions['C'].width = 40
    ws_testcases.column_dimensions['D'].width = 50
    ws_testcases.column_dimensions['E'].width = 35
    ws_testcases.column_dimensions['F'].width = 35
    ws_testcases.column_dimensions['G'].width = 10
    ws_testcases.column_dimensions['H'].width = 15
    ws_testcases.column_dimensions['I'].width = 15
    ws_testcases.column_dimensions['J'].width = 50

    print('✅ Đã tạo sheet Test cases')

    # Lưu file
    output_file = 'docs/Test_Cases_Converted.xlsx'
    wb_dest.save(output_file)
    
    print(f'\n🎉 CHUYỂN ĐỔI TEST CASES THÀNH CÔNG!')
    print(f'📁 File đầu ra: {output_file}')
    print(f'📊 Tổng số: {len(test_cases)} test cases trong {len(features)} features')
    
    # In mapping
    print('\n📋 Feature ID Mapping:')
    for feature_name, feature_id in feature_id_map.items():
        print(f'   {feature_id}: {feature_name}')
    
    return feature_id_map


def convert_bug_reports(feature_id_map=None):
    """Chuyển đổi Bug Reports"""
    # Đọc file nguồn
    source_file = 'functional_testing/reports/Bug_reports.xlsx'
    print(f'\n📖 Đang đọc file Bug nguồn: {source_file}')
    wb_source = openpyxl.load_workbook(source_file)
    ws_bugs = wb_source['Bug Reports']

    # Đọc template đích
    dest_template = 'docs/Bug_Report_Template.xlsx'
    print(f'📋 Đang đọc Bug template: {dest_template}')
    wb_dest = openpyxl.load_workbook(dest_template)

    # --- CHUẨN BỊ DỮ LIỆU ---
    # Cột file nguồn: Bug ID, Test Case ID, Summary, Description, Severity, Priority, 
    # Module, Feature, Steps to Reproduce, Expected Result, Actual Result, Status, 
    # Assignee, Found Date, Fixed Date, Verified By
    bugs = []
    for row in ws_bugs.iter_rows(min_row=2, values_only=True):
        if row[0]:
            bugs.append({
                'bug_id': row[0],           # Bug ID
                'tc_id': row[1],            # Test Case ID
                'summary': row[2],          # Summary
                'description': row[3],      # Description
                'severity': row[4],         # Severity
                'priority': row[5],         # Priority
                'module': row[6],           # Module
                'feature': row[7],          # Feature
                'steps': row[8],            # Steps to Reproduce
                'expected': row[9],         # Expected Result
                'actual': row[10],          # Actual Result
                'status': row[11],          # Status
                'assignee': row[12],        # Assignee
                'found_date': row[13],      # Found Date
                'fixed_date': row[14],      # Fixed Date
                'verified_by': row[15],     # Verified By
            })

    print(f'✅ Đọc được {len(bugs)} bugs')

    # --- XỬ LÝ SHEET Test summary report ---
    ws_summary = wb_dest['Test summary report']
    ws_summary['A1'] = 'Tester'
    ws_summary['B1'] = 'Lê Hoàng Việt'
    ws_summary['A2'] = 'Date'
    ws_summary['B2'] = datetime.now().strftime('%Y-%m-%d')

    # --- XỬ LÝ SHEET Bug report ---
    ws_bug_report = wb_dest['Bug report']

    # Styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Xóa dữ liệu mẫu cũ (giữ header row 1)
    ws_bug_report.delete_rows(2, ws_bug_report.max_row)

    # Thêm bugs theo format template
    # Template columns: Defect ID, Defect Title, Defect Description, Steps to Reproduce, 
    # Expected Result, Actual Result, Function ID, Severity, Reported By, Date Reported, Status, Comment
    current_row = 2
    for idx, bug in enumerate(bugs, 1):
        defect_id = f'B{idx:03d}'
        
        # Tìm Function ID từ feature_id_map nếu có
        function_id = ''
        if feature_id_map:
            feature_key = f"{bug['module']} - {bug['feature']}"
            function_id = feature_id_map.get(feature_key, bug['tc_id'])
        else:
            function_id = bug['tc_id']
        
        # Ghi dữ liệu
        ws_bug_report[f'A{current_row}'] = defect_id
        ws_bug_report[f'B{current_row}'] = bug['summary']
        ws_bug_report[f'C{current_row}'] = bug['description']
        ws_bug_report[f'D{current_row}'] = bug['steps']
        ws_bug_report[f'E{current_row}'] = bug['expected']
        ws_bug_report[f'F{current_row}'] = bug['actual']
        ws_bug_report[f'G{current_row}'] = function_id
        ws_bug_report[f'H{current_row}'] = bug['severity']
        ws_bug_report[f'I{current_row}'] = 'Lê Hoàng Việt'  # Reported By
        ws_bug_report[f'J{current_row}'] = bug['found_date']
        ws_bug_report[f'K{current_row}'] = bug['status']
        ws_bug_report[f'L{current_row}'] = f"TC: {bug['tc_id']} | Priority: {bug['priority']}"  # Comment
        
        # Apply styles
        for col in range(1, 13):
            cell = ws_bug_report.cell(row=current_row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
        
        current_row += 1

    # Điều chỉnh độ rộng cột
    ws_bug_report.column_dimensions['A'].width = 10
    ws_bug_report.column_dimensions['B'].width = 40
    ws_bug_report.column_dimensions['C'].width = 50
    ws_bug_report.column_dimensions['D'].width = 40
    ws_bug_report.column_dimensions['E'].width = 30
    ws_bug_report.column_dimensions['F'].width = 30
    ws_bug_report.column_dimensions['G'].width = 12
    ws_bug_report.column_dimensions['H'].width = 10
    ws_bug_report.column_dimensions['I'].width = 15
    ws_bug_report.column_dimensions['J'].width = 15
    ws_bug_report.column_dimensions['K'].width = 10
    ws_bug_report.column_dimensions['L'].width = 30

    print('✅ Đã tạo sheet Bug report')

    # Lưu file
    output_file = 'docs/Bug_Report_Converted.xlsx'
    wb_dest.save(output_file)
    
    print(f'\n🎉 CHUYỂN ĐỔI BUG REPORTS THÀNH CÔNG!')
    print(f'📁 File đầu ra: {output_file}')
    print(f'🐛 Tổng số: {len(bugs)} bugs')


if __name__ == '__main__':
    print('=' * 60)
    print('🔄 BẮT ĐẦU CHUYỂN ĐỔI DỮ LIỆU')
    print('=' * 60)
    
    # Chuyển đổi Test Cases trước
    feature_id_map = convert_test_cases()
    
    print('\n' + '-' * 60)
    
    # Chuyển đổi Bug Reports
    convert_bug_reports(feature_id_map)
    
    print('\n' + '=' * 60)
    print('✅ HOÀN TẤT TẤT CẢ CHUYỂN ĐỔI!')
    print('=' * 60)
